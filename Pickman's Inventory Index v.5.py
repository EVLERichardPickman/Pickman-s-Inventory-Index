import os
import json
import sys
import requests
import ctypes
import ctypes.wintypes as wintypes

from PyQt5.QtCore import Qt, QEvent
from PyQt5.QtGui import QIcon, QPalette, QColor, QBrush, QTextDocument
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLineEdit,
    QPushButton,
    QLabel,
    QTableWidget,
    QTableWidgetItem,
    QMessageBox,
    QComboBox,
    QHeaderView,
    QFileDialog,
    QSizePolicy,
)
from PyQt5.QtPrintSupport import QPrinter

from openpyxl import Workbook, load_workbook


# ---------- PATH HELPERS FOR PORTABLE .EXE ----------

def get_app_dir():
    """
    Return the directory where the script/.exe lives.
    - When running from source: folder of this .py file
    - When built as an .exe with PyInstaller: folder of the .exe
    """
    if getattr(sys, "frozen", False):  # PyInstaller sets this attribute
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


APP_DIR = get_app_dir()
INVENTORY_FILE = os.path.join(APP_DIR, "inventory.json")   # saved next to .py/.exe
SETTINGS_FILE = os.path.join(APP_DIR, "settings.json")     # theme persistence


# ---------- EXE ICON HELPER (USE EMBEDDED ICON) ----------

def get_exe_icon_qicon():
    """
    On Windows, load the first icon resource from the running executable
    and return it as a QIcon. Returns None on failure or non-Windows.

    This lets a single-file PyInstaller .exe use its own embedded icon
    for the Qt window/taskbar icon, without needing an external .ico file.
    """
    if not sys.platform.startswith("win"):
        return None

    try:
        # When frozen by PyInstaller, sys.executable is the .exe
        exe_path = sys.executable if getattr(sys, "frozen", False) else os.path.abspath(sys.argv[0])

        shell32 = ctypes.windll.shell32
        user32 = ctypes.windll.user32

        # Prepare handles for large and small icons
        hicon_large = wintypes.HICON()
        hicon_small = wintypes.HICON()

        # Extract first icon (index 0)
        num_icons = shell32.ExtractIconExW(
            exe_path,
            0,
            ctypes.byref(hicon_large),
            ctypes.byref(hicon_small),
            1
        )

        if num_icons <= 0:
            return None

        # Prefer small icon if available, otherwise large
        hicon = hicon_small if hicon_small.value else hicon_large
        if not hicon or not hicon.value:
            return None

        from PyQt5.QtGui import QPixmap, QIcon as _QIcon

        pixmap = QPixmap.fromWinHICON(hicon.value)
        # Free the Windows icon handle
        user32.DestroyIcon(hicon.value)

        if pixmap.isNull():
            return None

        return _QIcon(pixmap)
    except Exception:
        return None


# --- API ENDPOINTS ---
UEX_API_URL = "https://api.uexcorp.uk/2.0/marketplace_averages_all"
UEX_CATEGORIES_URL = "https://api.uexcorp.uk/2.0/categories"


# ---------- SORTING-HELPER ITEM ----------

class NumericTableWidgetItem(QTableWidgetItem):
    """
    A table item that tries to sort numerically if possible.
    - For Qty, Listed Price, Sell Price, Line Total columns (numbers with commas or blanks)
    - Falls back to plain string comparison for non-numeric values
    """
    def _to_number(self, text):
        text = text.replace(",", "").strip()
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            # If it's not numeric, return None and let caller decide
            return None

    def __lt__(self, other):
        if not isinstance(other, QTableWidgetItem):
            return super().__lt__(other)

        self_num = self._to_number(self.text())
        other_num = self._to_number(other.text())

        # If both are numeric, compare numerically
        if self_num is not None and other_num is not None:
            return self_num < other_num

        # Otherwise, fall back to string comparison (useful for Item Name)
        return self.text().lower() < other.text().lower()


def fetch_market_data():
    """
    Fetch marketplace 30-day averages for all items from UEX.
    Returns a list of dicts or raises an exception on error.
    """
    resp = requests.get(UEX_API_URL, timeout=15)
    resp.raise_for_status()
    data = resp.json()

    if not isinstance(data, dict) or data.get("status") != "ok":
        raise RuntimeError(f"Unexpected API response: {data!r}")

    return data.get("data", [])


def fetch_categories():
    """
    Fetch item/service/contract categories and return a mapping:
        { id_category (int): {"section": str, "name": str} }
    We only care about type == "item".
    """
    resp = requests.get(UEX_CATEGORIES_URL, timeout=15)
    resp.raise_for_status()
    data = resp.json()

    if not isinstance(data, dict) or data.get("status") != "ok":
        raise RuntimeError(f"Unexpected categories API response: {data!r}")

    categories_map = {}
    for cat in data.get("data", []):
        # Only item categories are relevant here
        if cat.get("type") and cat.get("type") != "item":
            continue

        cat_id = cat.get("id")
        if cat_id is None:
            continue

        try:
            cat_id = int(cat_id)
        except (TypeError, ValueError):
            continue

        categories_map[cat_id] = {
            "section": cat.get("section") or "",
            "name": cat.get("name") or "",
        }

    return categories_map


def load_inventory(path):
    """
    Load saved inventory from JSON.

    Supports two formats (backwards compatible):
      - { item_key: number }                         # old style: quantity only
      - { item_key: {"qty": number, "sell_price": number_or_str} }  # new style
    """
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_inventory(path, inventory_dict):
    """
    Save inventory (qty + sell_price) to JSON.
    """
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(inventory_dict, f, indent=2)
    except Exception as e:
        print(f"Failed to save inventory: {e}")


# ---------- SETTINGS (THEME PERSISTENCE) ----------

def load_settings(path):
    """
    Load app settings (e.g., theme) from JSON.
    """
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_settings(path, settings_dict):
    """
    Save app settings (e.g., theme) to JSON.
    """
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(settings_dict, f, indent=2)
    except Exception as e:
        print(f"Failed to save settings: {e}")


class AnnexWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # Use the exe's embedded icon if available
        icon = get_exe_icon_qicon()
        if icon is not None:
            self.setWindowIcon(icon)

        # Window title + size
        self.setWindowTitle("Pickman's Inventory Index")

        # Data
        self.market_data = []         # full list from API
        self.filtered_data = []       # what is currently displayed
        self.inventory = load_inventory(INVENTORY_FILE)  # {item_key: qty or {qty, sell_price}}
        self.categories = {}          # {id_category: {"section": ..., "name": ...}}
        self.category_map = {}        # {"Armor": set(["Arms", "Backpacks", ...]), ...}
        self.updating_table = False   # guard for itemChanged recursion

        self._central_widget = None   # for wheel event routing

        # Theme settings (load last used)
        self.settings = load_settings(SETTINGS_FILE)
        if not isinstance(self.settings, dict):
            self.settings = {}
        theme = self.settings.get("theme", "dark")
        self.dark_mode = (theme != "light")   # default to dark if missing/invalid
        self.theme_button = None

        # UI
        self._build_ui()

        # Apply theme based on saved setting
        if self.dark_mode:
            self.apply_dark_theme()
        else:
            self.apply_light_theme()

        # Load data from UEX
        self.load_data()

    # ---------------- THEME HANDLING ----------------
    def apply_dark_theme(self):
        app = QApplication.instance()
        if not app:
            return

        palette = QPalette()
        # Dark background, light text
        palette.setColor(QPalette.Window, QColor(30, 30, 30))
        palette.setColor(QPalette.WindowText, Qt.white)
        palette.setColor(QPalette.Base, QColor(20, 20, 20))
        palette.setColor(QPalette.AlternateBase, QColor(45, 45, 45))
        palette.setColor(QPalette.ToolTipBase, Qt.white)
        palette.setColor(QPalette.ToolTipText, Qt.white)
        palette.setColor(QPalette.Text, Qt.white)
        palette.setColor(QPalette.Button, QColor(45, 45, 45))
        palette.setColor(QPalette.ButtonText, Qt.white)
        palette.setColor(QPalette.BrightText, Qt.red)
        palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        palette.setColor(QPalette.HighlightedText, Qt.black)

        app.setPalette(palette)

        # Dark styles for widgets, buttons, filters, and table
        app.setStyleSheet(
            """
            QWidget {
                background-color: #1e1e1e;
                color: #f0f0f0;
            }

            QLineEdit, QComboBox {
                background-color: #252525;
                color: #f0f0f0;
                border: 1px solid #555555;
            }

            QComboBox QAbstractItemView {
                background-color: #252525;
                color: #f0f0f0;
                selection-background-color: #2a82da;
                selection-color: #000000;
            }

            QTableWidget {
                background-color: #252525;
                color: #f0f0f0;
                gridline-color: #555555;
                selection-background-color: #2a82da;
                selection-color: #000000;
            }

            QHeaderView::section {
                background-color: #303030;
                color: #f0f0f0;
                border: 1px solid #555555;
            }

            QPushButton {
                background-color: #3b3b3b;
                color: #f0f0f0;
                border: 1px solid #666666;
                padding: 3px 8px;
            }

            QPushButton:hover {
                background-color: #4a4a4a;
            }

            QPushButton:pressed {
                background-color: #2f2f2f;
            }
            """
        )

        self.dark_mode = True
        if self.theme_button:
            # When in dark mode, button offers to switch to Light
            self.theme_button.setText("Light")

    def apply_light_theme(self):
        app = QApplication.instance()
        if not app:
            return

        palette = QPalette()
        # Light background, dark text (inverse of dark)
        palette.setColor(QPalette.Window, Qt.white)
        palette.setColor(QPalette.WindowText, Qt.black)
        palette.setColor(QPalette.Base, Qt.white)
        palette.setColor(QPalette.AlternateBase, QColor(235, 235, 235))
        palette.setColor(QPalette.ToolTipBase, Qt.black)
        palette.setColor(QPalette.ToolTipText, Qt.white)
        palette.setColor(QPalette.Text, Qt.black)
        palette.setColor(QPalette.Button, QColor(235, 235, 235))
        palette.setColor(QPalette.ButtonText, Qt.black)
        palette.setColor(QPalette.BrightText, Qt.red)
        palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
        palette.setColor(QPalette.HighlightedText, Qt.white)

        app.setPalette(palette)

        # Light styles: inverse of dark theme
        app.setStyleSheet(
            """
            QWidget {
                background-color: #f0f0f0;
                color: #202020;
            }

            QLineEdit, QComboBox {
                background-color: #ffffff;
                color: #202020;
                border: 1px solid #b0b0b0;
            }

            QComboBox QAbstractItemView {
                background-color: #ffffff;
                color: #202020;
                selection-background-color: #2a82da;
                selection-color: #ffffff;
            }

            QTableWidget {
                background-color: #ffffff;
                color: #202020;
                gridline-color: #b0b0b0;
                selection-background-color: #2a82da;
                selection-color: #ffffff;
            }

            QHeaderView::section {
                background-color: #e0e0e0;
                color: #202020;
                border: 1px solid #b0b0b0;
            }

            QPushButton {
                background-color: #e0e0e0;
                color: #202020;
                border: 1px solid #a0a0a0;
                padding: 3px 8px;
            }

            QPushButton:hover {
                background-color: #d0d0d0;
            }

            QPushButton:pressed {
                background-color: #c0c0c0;
            }
            """
        )

        self.dark_mode = False
        if self.theme_button:
            # When in light mode, button offers to switch to Dark
            self.theme_button.setText("Dark")

    def toggle_theme(self):
        if self.dark_mode:
            self.apply_light_theme()
        else:
            self.apply_dark_theme()

        # Persist theme to settings.json
        try:
            if not isinstance(self.settings, dict):
                self.settings = {}
            self.settings["theme"] = "dark" if self.dark_mode else "light"
            save_settings(SETTINGS_FILE, self.settings)
        except Exception as e:
            print(f"Failed to save theme setting: {e}")

    # ---------- AUTOSIZE HELPERS ----------
    def autosize_button(self, button: QPushButton):
        """
        Make the button wide enough for its text and prevent it
        from being shrunk below that width.
        """
        hint = button.sizeHint()
        min_w = hint.width() + 6
        min_h = hint.height()
        button.setMinimumSize(min_w, min_h)

        sp = button.sizePolicy()
        sp.setHorizontalPolicy(QSizePolicy.Fixed)
        sp.setVerticalPolicy(QSizePolicy.Fixed)
        button.setSizePolicy(sp)

    def autosize_combo(self, combo: QComboBox):
        """
        Ensure the combo box is wide enough for its longest item
        and does not get shrunk smaller than that.
        """
        combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        combo.setMinimumContentsLength(0)

        hint = combo.sizeHint()
        combo.setMinimumWidth(hint.width() + 6)

        sp = combo.sizePolicy()
        sp.setHorizontalPolicy(QSizePolicy.Fixed)
        sp.setVerticalPolicy(QSizePolicy.Fixed)
        combo.setSizePolicy(sp)

    # ---------------- UI BUILD ----------------
    def _build_ui(self):
        central = QWidget()
        self._central_widget = central
        self.setCentralWidget(central)

        # Route wheel events from central widget to the table
        self._central_widget.installEventFilter(self)

        main_layout = QVBoxLayout()
        central.setLayout(main_layout)

        # Row: basic filter (all / inventory)
        filter_layout = QHBoxLayout()
        main_layout.addLayout(filter_layout)

        filter_layout.addWidget(QLabel("Filter:"))
        self.filter_combo = QComboBox()
        self.filter_combo.addItems([
            "UEX",
            "Inventory",
        ])
        # Default to "Inventory only" on startup
        self.filter_combo.setCurrentIndex(1)
        self.filter_combo.currentIndexChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.filter_combo)

        filter_layout.addStretch()

        # Theme toggle button (top-right)
        self.theme_button = QPushButton("Light" if self.dark_mode else "Dark")
        self.theme_button.clicked.connect(self.toggle_theme)
        filter_layout.addWidget(self.theme_button)

        # Row: Category / Subcategory filters (Armor -> Arms, etc.)
        category_layout = QHBoxLayout()
        main_layout.addLayout(category_layout)

        category_layout.addWidget(QLabel("Category:"))
        self.category_combo = QComboBox()
        self.category_combo.addItem("All Categories", None)
        self.category_combo.currentIndexChanged.connect(self.on_category_changed)
        category_layout.addWidget(self.category_combo)

        category_layout.addWidget(QLabel("Subcategory:"))
        self.subcategory_combo = QComboBox()
        self.subcategory_combo.addItem("All Subcategories", None)
        self.subcategory_combo.currentIndexChanged.connect(self.apply_filters)
        category_layout.addWidget(self.subcategory_combo)

        category_layout.addStretch()

        # Top search row
        search_layout = QHBoxLayout()
        main_layout.addLayout(search_layout)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText(
            "Search by partial name (space or comma separated)…"
        )
        # Live search as you type
        self.search_edit.textChanged.connect(self.apply_filters)
        self.search_edit.returnPressed.connect(self.apply_filters)
        search_layout.addWidget(self.search_edit)

        # Home button only
        home_button = QPushButton("Home")
        home_button.clicked.connect(self.go_home)
        search_layout.addWidget(home_button)

        self.status_label = QLabel("Loading data from UEX…")
        self.status_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        search_layout.addWidget(self.status_label)

        # Table
        self.table = QTableWidget()
        # Qty, Item Name, Listed Price, Trend, Sell Price, Line Total
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(
            [
                "Qty",
                "Item Name",
                "Listed Price (aUEC)",
                "Trend",
                "Sell Price (aUEC)",
                "Line Total (aUEC)",
            ]
        )

        # Make "Item Name" the largest/stretch column
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # Qty
        header.setSectionResizeMode(1, QHeaderView.Stretch)           # Item Name
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Listed Price
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Trend
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Sell Price
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # Line Total

        # Enable click-to-sort on headers
        self.table.setSortingEnabled(True)

        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QTableWidget.AllEditTriggers)
        self.table.itemChanged.connect(self.on_item_changed)
        main_layout.addWidget(self.table)

        # Bottom area: buttons + totals
        bottom_layout = QVBoxLayout()
        main_layout.addLayout(bottom_layout)

        # --- Row 1: Import Index + Sell Total ---
        row1 = QHBoxLayout()
        bottom_layout.addLayout(row1)

        self.import_button = QPushButton("Import Index")
        self.import_button.clicked.connect(self.import_index)
        row1.addWidget(self.import_button)

        row1.addStretch()
        row1.addWidget(QLabel("Sell Total:"))

        self.total_label = QLabel("0")
        font = self.total_label.font()
        font.setBold(True)
        self.total_label.setFont(font)
        row1.addWidget(self.total_label)

        # --- Row 2: Export Index + Overall Inventory Value ---
        row2 = QHBoxLayout()
        bottom_layout.addLayout(row2)

        self.export_button = QPushButton("Export Index")
        self.export_button.clicked.connect(self.export_index)
        row2.addWidget(self.export_button)

        row2.addStretch()
        row2.addWidget(QLabel("Overall Inventory Value:"))

        self.overall_total_label = QLabel("0")
        top_font = self.overall_total_label.font()
        top_font.setBold(True)
        self.overall_total_label.setFont(top_font)
        row2.addWidget(self.overall_total_label)

        # ---- AUTOSIZE ALL BUTTONS + COMBOS AFTER CREATION ----
        # Combo boxes
        self.autosize_combo(self.filter_combo)
        self.autosize_combo(self.category_combo)
        self.autosize_combo(self.subcategory_combo)

        # Buttons
        self.autosize_button(self.theme_button)
        self.autosize_button(home_button)
        self.autosize_button(self.import_button)
        self.autosize_button(self.export_button)

    # ---------------- EVENT FILTER (MOUSE WHEEL) ----------------
    def eventFilter(self, obj, event):
        # If the user scrolls the wheel anywhere on the central widget,
        # forward the wheel event to the table so it scrolls.
        if obj is self._central_widget and event.type() == QEvent.Wheel:
            QApplication.sendEvent(self.table.viewport(), event)
            return True
        return super().eventFilter(obj, event)

    # ---------------- DATA LOADING ----------------
    def load_data(self):
        try:
            self.status_label.setText("Contacting UEX API…")
            QApplication.processEvents()

            # Prices
            self.market_data = fetch_market_data()

            # Categories (Armor, Avionics, etc.) and subcategories (Arms, Backpacks, etc.)
            try:
                self.categories = fetch_categories()
            except Exception as e:
                print(f"Warning: failed to load categories: {e}")
                self.categories = {}

            # Attach category labels to each item using id_category
            for item in self.market_data:
                cat_id = item.get("id_category")
                try:
                    cat_id_int = int(cat_id) if cat_id is not None else None
                except (TypeError, ValueError):
                    cat_id_int = None

                cat_info = self.categories.get(cat_id_int)
                if cat_info:
                    item["category_section"] = (cat_info.get("section") or "").strip()
                    item["category_name"] = (cat_info.get("name") or "").strip()
                else:
                    item["category_section"] = ""
                    item["category_name"] = ""

            # ---------- TREND ARROWS (price_sell vs price_buy) ----------
            # UEX gives 30-day average buy/sell. We compare those:
            #   ▲ if price_sell > price_buy
            #   ▼ if price_sell < price_buy
            #   → otherwise (equal, or missing data)
            for item in self.market_data:
                price_sell = item.get("price_sell")
                price_buy = item.get("price_buy")

                try:
                    ps = float(price_sell) if price_sell is not None else None
                except (TypeError, ValueError):
                    ps = None

                try:
                    pb = float(price_buy) if price_buy is not None else None
                except (TypeError, ValueError):
                    pb = None

                trend_arrow = "→"
                if ps is not None and pb is not None:
                    if ps > pb:
                        trend_arrow = "▲"
                    elif ps < pb:
                        trend_arrow = "▼"

                item["trend_arrow"] = trend_arrow

            # Build the category / subcategory dropdowns
            self.populate_category_filters()

            self.status_label.setText(f"Loaded {len(self.market_data)} items from UEX.")
            # Apply filters (which also alphabetizes). Default filter is Inventory-only.
            self.apply_filters()
            self.update_overall_total()
        except Exception as e:
            self.status_label.setText("Failed to load data from UEX.")
            QMessageBox.critical(self, "Error", f"Error fetching UEX data:\n{e}")

    # ---------------- CATEGORY FILTER SETUP ----------------
    def populate_category_filters(self):
        """
        Build:
          - self.category_map: {section: set(subcategory_names)}
          - fill category_combo and subcategory_combo.
        """
        self.category_map = {}
        for item in self.market_data:
            section = (item.get("category_section") or "").strip()
            name = (item.get("category_name") or "").strip()
            if not section:
                continue
            self.category_map.setdefault(section, set())
            if name:
                self.category_map[section].add(name)

        # Populate Category combo
        self.category_combo.blockSignals(True)
        self.category_combo.clear()
        self.category_combo.addItem("All Categories", None)
        for section in sorted(self.category_map.keys()):
            self.category_combo.addItem(section, section)
        self.category_combo.blockSignals(False)

        # Populate Subcategory combo for the initial selection
        self.populate_subcategories()

    def populate_subcategories(self):
        """
        Fill subcategory combo based on current category selection.
        """
        self.subcategory_combo.blockSignals(True)
        self.subcategory_combo.clear()
        self.subcategory_combo.addItem("All Subcategories", None)

        selected_section = self.category_combo.currentData()

        if selected_section is None:
            # All categories: show union of all subcategories
            all_subs = set()
            for subs in self.category_map.values():
                all_subs.update(subs)
            for name in sorted(all_subs):
                self.subcategory_combo.addItem(name, name)
        else:
            for name in sorted(self.category_map.get(selected_section, [])):
                self.subcategory_combo.addItem(name, name)

        self.subcategory_combo.blockSignals(False)

    def on_category_changed(self, index):
        """
        When the Category combo changes, update subcategories and re-apply filters.
        """
        self.populate_subcategories()
        self.apply_filters()

    # ---------------- HELPERS ----------------
    def get_unit_price(self, item):
        """
        Decide Listed Price for an item: prefer price_sell, fallback to price_buy.
        """
        price_sell = item.get("price_sell")
        price_buy = item.get("price_buy")
        if price_sell is not None:
            return float(price_sell)
        if price_buy is not None:
            return float(price_buy)
        return 0.0

    def item_key(self, item):
        """
        Unique-ish key per item for inventory persistence.
        Prefer UUID; fallback to id_item; then item_name.
        """
        if item.get("item_uuid"):
            return f"uuid:{item['item_uuid']}"
        elif item.get("id_item") is not None:
            return f"id:{item['id_item']}"
        else:
            return f"name:{item.get('item_name', 'unknown')}"

    def get_qty(self, key):
        """
        Get quantity for an item key from self.inventory, handling both old and new formats.
        """
        val = self.inventory.get(key)
        if isinstance(val, dict):
            val = val.get("qty", 0)
        if val is None:
            return 0.0
        try:
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    def get_sell_price_for_key(self, key):
        """
        Get stored Sell Price for an item key from self.inventory.
        Returns:
          - float if numeric
          - string if non-numeric
          - "" if not set
        """
        val = self.inventory.get(key)
        if not isinstance(val, dict):
            return ""
        sp = val.get("sell_price")
        if sp in (None, ""):
            return ""
        try:
            return float(sp)
        except (TypeError, ValueError):
            return sp

    # ---------------- FILTERING + SEARCH ----------------
    def apply_filters(self):
        """
        Apply:
          - partial-name search
          - inventory filter (all / inventory)
          - category & subcategory filters
        Then alphabetize result by item_name.
        """
        if not self.market_data:
            self.filtered_data = []
            self.populate_table()
            return

        raw = self.search_edit.text().strip()
        keywords = [k.lower() for k in raw.replace(",", " ").split() if k.strip()]
        filter_mode = self.filter_combo.currentIndex()  # 0 = all, 1 = inventory

        cat_filter = self.category_combo.currentData()
        subcat_filter = self.subcategory_combo.currentData()

        filtered = []

        for item in self.market_data:
            name = (item.get("item_name") or "").lower()

            # Partial-name search: match if ANY keyword appears in name.
            if keywords and not any(k in name for k in keywords):
                continue

            # Inventory filter
            key = self.item_key(item)
            qty = self.get_qty(key)

            if filter_mode == 1 and qty <= 0:
                # Inventory only
                continue

            # Category / Subcategory filters
            section = (item.get("category_section") or "").strip()
            subname = (item.get("category_name") or "").strip()

            if cat_filter is not None and section != cat_filter:
                continue
            if subcat_filter is not None and subname != subcat_filter:
                continue

            filtered.append(item)

        # Default order: alphabetize by item name
        filtered.sort(key=lambda it: (it.get("item_name") or "").lower())
        self.filtered_data = filtered

        # Status text
        parts = []
        if keywords:
            parts.append("search")
        if filter_mode != 0:
            parts.append("inv-filter")
        if cat_filter is not None or subcat_filter is not None:
            parts.append("category-filter")

        if parts:
            self.status_label.setText(
                f"Showing {len(self.filtered_data)} items ({', '.join(parts)})."
            )
        else:
            self.status_label.setText(f"Showing {len(self.filtered_data)} items.")

        self.populate_table()

    def go_home(self):
        """
        Reset to default home view:
          - inventory only
          - clear search
          - all categories, all subcategories
        """
        # Block signals so we don't trigger apply_filters multiple times mid-reset
        self.search_edit.blockSignals(True)
        self.filter_combo.blockSignals(True)
        self.category_combo.blockSignals(True)
        self.subcategory_combo.blockSignals(True)

        self.search_edit.clear()
        self.filter_combo.setCurrentIndex(1)   # Inventory only
        self.category_combo.setCurrentIndex(0) # All Categories
        self.populate_subcategories()
        self.subcategory_combo.setCurrentIndex(0)  # All Subcategories

        # Re-enable signals
        self.search_edit.blockSignals(False)
        self.filter_combo.blockSignals(False)
        self.category_combo.blockSignals(False)
        self.subcategory_combo.blockSignals(False)

        # Apply filters once with the "home" state
        self.apply_filters()

    # ---------------- TABLE POPULATION ----------------
    def populate_table(self):
        # Guard so on_item_changed ignores any signals triggered here
        self.updating_table = True

        # Temporarily disable sorting and block itemChanged while we repopulate rows
        sorting_was_enabled = self.table.isSortingEnabled()
        self.table.setSortingEnabled(False)
        self.table.blockSignals(True)

        self.table.clearContents()
        self.table.setRowCount(len(self.filtered_data))

        for row, item in enumerate(self.filtered_data):
            key = self.item_key(item)
            unit_price = self.get_unit_price(item)
            trend_arrow = item.get("trend_arrow", "→")

            # Quantity (from inventory or 0)
            qty_float = self.get_qty(key)

            # --- Qty display with reduced resolution (e.g. 17.0 -> 17, 0 -> "") ---
            if qty_float == 0:
                qty_display = ""
            elif float(qty_float).is_integer():
                qty_display = str(int(qty_float))
            else:
                qty_display = str(qty_float)

            # Qty (numeric-sort)
            qty_item = NumericTableWidgetItem(qty_display)
            qty_item.setTextAlignment(Qt.AlignCenter)
            # Store key in user data so we can find this row even after sorting
            qty_item.setData(Qt.UserRole, key)
            self.table.setItem(row, 0, qty_item)

            # Item name (string-sort)
            name = item.get("item_name") or f"Item {item.get('id_item')}"
            name_item = NumericTableWidgetItem(name)
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 1, name_item)

            # Listed price (numeric-sort, read-only)
            price_text = f"{unit_price:,.0f}"
            price_item = NumericTableWidgetItem(price_text)
            price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            price_item.setFlags(price_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 2, price_item)

            # Trend arrow (read-only, colored)
            trend_item = QTableWidgetItem(trend_arrow)
            trend_item.setTextAlignment(Qt.AlignCenter)
            trend_item.setFlags(trend_item.flags() & ~Qt.ItemIsEditable)

            # Color arrows: green up, red down, default for →
            if trend_arrow == "▲":
                trend_item.setForeground(QBrush(QColor(0, 200, 0)))   # green
            elif trend_arrow == "▼":
                trend_item.setForeground(QBrush(QColor(200, 0, 0)))   # red

            self.table.setItem(row, 3, trend_item)

            # Sell price (numeric-sort, EDITABLE) — loaded from inventory if present
            sell_val = self.get_sell_price_for_key(key)

            if isinstance(sell_val, (int, float)):
                # For money: blank if 0 or negative, otherwise format with commas
                if sell_val <= 0:
                    sell_display = ""
                else:
                    sell_display = f"{sell_val:,.0f}"
            else:
                # Non-numeric or empty value
                if sell_val in ("", None):
                    sell_display = ""
                else:
                    sell_display = str(sell_val)

            sell_item = NumericTableWidgetItem(sell_display)
            sell_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row, 4, sell_item)

            # Line total (numeric-sort, read-only) based on listed price * qty
            line_total = qty_float * unit_price
            line_text = f"{line_total:,.0f}" if line_total else ""
            line_item = NumericTableWidgetItem(line_text)
            line_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            line_item.setFlags(line_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 5, line_item)

        # Re-enable signals and restore sorting state
        self.table.blockSignals(False)
        self.table.setSortingEnabled(sorting_was_enabled)

        self.updating_table = False
        self.update_grand_total()

    # ---------------- TABLE CHANGE HANDLING ----------------
    def on_item_changed(self, item):
        if self.updating_table:
            return

        row = item.row()
        col = item.column()

        if row < 0 or row >= self.table.rowCount():
            return

        # We always use the Qty cell to recover the key
        qty_item = self.table.item(row, 0)
        if qty_item is None:
            return
        key = qty_item.data(Qt.UserRole)
        if not key:
            return

        # ----- Qty changed -----
        if col == 0:
            text = item.text().replace(",", "").strip()
            try:
                qty = float(text) if text else 0.0
            except ValueError:
                qty = 0.0

            # Read existing record if present
            rec = self.inventory.get(key)
            if isinstance(rec, dict):
                rec = dict(rec)  # shallow copy
            else:
                rec = {}

            rec["qty"] = qty

            if qty == 0:
                # If qty is zero and no sell_price is stored, remove the entry
                if "sell_price" not in rec or rec["sell_price"] in ("", None):
                    self.inventory.pop(key, None)
                else:
                    # Keep entry with qty=0 because sell_price is still meaningful
                    self.inventory[key] = rec
            else:
                self.inventory[key] = rec

            save_inventory(INVENTORY_FILE, self.inventory)

            # Recalc this row + both totals
            self.recalculate_row(row)
            self.update_grand_total()
            self.update_overall_total()
            return

        # ----- Sell Price changed -----
        if col == 4:
            # Raw text, but strip commas for numeric parsing
            raw_text = item.text()
            text = raw_text.replace(",", "").strip()

            # Read existing record
            rec = self.inventory.get(key)
            if isinstance(rec, dict):
                rec = dict(rec)
            else:
                # If we had old-style numeric qty only, convert it
                rec = {"qty": self.get_qty(key)}

            display_text = ""

            if not text:
                # User cleared the field: remove sell_price
                rec.pop("sell_price", None)
            else:
                # Try numeric; if fails, treat as free-form string
                try:
                    sp = float(text)
                    # For money: if 0 or negative, treat as empty / no sell price
                    if sp <= 0:
                        rec.pop("sell_price", None)
                        display_text = ""
                    else:
                        rec["sell_price"] = sp
                        # Format with thousands separators, no decimals
                        display_text = f"{sp:,.0f}"
                except ValueError:
                    # Non-numeric input: store and display as-is
                    rec["sell_price"] = raw_text
                    display_text = raw_text

            # If qty == 0 and no sell_price, remove entry entirely
            qty_val = rec.get("qty", 0)
            try:
                qty_val = float(qty_val)
            except (TypeError, ValueError):
                qty_val = 0.0

            if qty_val == 0 and "sell_price" not in rec:
                self.inventory.pop(key, None)
            else:
                self.inventory[key] = rec

            save_inventory(INVENTORY_FILE, self.inventory)

            # Re-format the cell text to match our display rules
            self.updating_table = True
            item.setText(display_text)
            self.updating_table = False

            # Sell total depends on sell price, so recalc it
            self.update_grand_total()
            return

    def recalculate_row(self, row):
        """
        Recalculate line total for the given row based on Qty and Listed Price.
        Also normalizes Qty display so 17.0 -> 17, 0 -> "".
        """
        self.updating_table = True

        qty_item = self.table.item(row, 0)
        price_item = self.table.item(row, 2)  # Listed price column
        line_item = self.table.item(row, 5)   # Line total column

        if qty_item is None or price_item is None or line_item is None:
            self.updating_table = False
            return

        qty_text = qty_item.text().replace(",", "").strip()
        price_text = price_item.text().replace(",", "").strip()

        try:
            qty = float(qty_text) if qty_text else 0.0
        except ValueError:
            qty = 0.0

        try:
            price = float(price_text) if price_text else 0.0
        except ValueError:
            price = 0.0

        # Normalize qty display
        if qty == 0:
            qty_display = ""
        elif float(qty).is_integer():
            qty_display = str(int(qty))
        else:
            qty_display = str(qty)
        qty_item.setText(qty_display)

        # Line total (still based on listed price)
        line_total = qty * price
        line_item.setText(f"{line_total:,.0f}" if line_total else "")

        self.updating_table = False

    # ---------------- TOTALS ----------------
    def update_grand_total(self):
        """
        Sell total of currently visible rows in the table (respects search/filter).
        Uses Qty * Sell Price (column 0 * column 4) for each row.
        Rows with blank/non-numeric/zero sell price are ignored.
        """
        total = 0.0
        rows = self.table.rowCount()

        for row in range(rows):
            qty_item = self.table.item(row, 0)   # Qty
            sell_item = self.table.item(row, 4)  # Sell Price
            if not qty_item or not sell_item:
                continue

            qty_text = qty_item.text().replace(",", "").strip()
            sell_text = sell_item.text().replace(",", "").strip()

            if not qty_text or not sell_text:
                continue

            try:
                qty = float(qty_text)
                sell_price = float(sell_text)
            except ValueError:
                # Non-numeric sell price (e.g., string labels) are ignored
                continue

            if qty <= 0 or sell_price <= 0:
                continue

            total += qty * sell_price

        self.total_label.setText(f"{total:,.0f}")

    def update_overall_total(self):
        """
        Overall inventory value:
        sum over ALL items (market_data) of qty * listed_price,
        using saved inventory quantities (even if not currently visible).
        """
        total = 0.0
        if not self.market_data:
            self.overall_total_label.setText("0")
            return

        for item in self.market_data:
            key = self.item_key(item)
            qty = self.get_qty(key)
            if qty <= 0:
                continue

            price = self.get_unit_price(item)
            total += qty * price

        self.overall_total_label.setText(f"{total:,.0f}")

    # ---------------- EXPORT INDEX (JSON, XLSX, TXT, PDF) ----------------
    def export_index(self):
        """
        Export the entire index (inventory) with fields:
        QTY, Item Name, Sell Value, Category, Sub-Category
        Formats: JSON, XLSX, TXT, PDF (based on chosen extension).
        """
        default_path = os.path.join(APP_DIR, "inventory_index")
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Export Index",
            default_path,
            "Index Files (*.json *.xlsx *.txt *.pdf);;JSON (*.json);;Excel (*.xlsx);;Text (*.txt);;PDF (*.pdf)"
        )
        if not file_path:
            return  # user cancelled

        # Ensure we have an extension
        root, ext = os.path.splitext(file_path)
        ext = ext.lower()
        if not ext:
            # Default to JSON if no extension given
            ext = ".json"
            file_path = root + ext

        if ext not in (".json", ".xlsx", ".txt", ".pdf"):
            QMessageBox.warning(
                self,
                "Unsupported Format",
                "Please use one of the supported extensions: .json, .xlsx, .txt, .pdf"
            )
            return

        # Build records from full inventory
        records = []
        for item in self.market_data:
            key = self.item_key(item)
            qty = self.get_qty(key)
            sell_price = self.get_sell_price_for_key(key)
            # Skip completely empty entries
            if qty <= 0 and (sell_price in ("", None)):
                continue

            name = item.get("item_name") or f"Item {item.get('id_item')}"
            category = item.get("category_section") or ""
            subcategory = item.get("category_name") or ""
            records.append({
                "QTY": qty,
                "Item Name": name,
                "Sell Value": sell_price if sell_price != "" else None,
                "Category": category,
                "Sub-Category": subcategory,
            })

        try:
            if ext == ".json":
                with open(file_path, "w", encoding="utf-8") as f:
                    json.dump(records, f, indent=2)

            elif ext == ".xlsx":
                wb = Workbook()
                ws = wb.active
                ws.title = "Index"
                ws.append(["QTY", "Item Name", "Sell Value", "Category", "Sub-Category"])
                for rec in records:
                    ws.append([
                        rec.get("QTY", 0),
                        rec.get("Item Name", ""),
                        rec.get("Sell Value", None),
                        rec.get("Category", ""),
                        rec.get("Sub-Category", ""),
                    ])
                wb.save(file_path)

            elif ext == ".txt":
                # Simple tab-separated text
                lines = ["QTY\tItem Name\tSell Value\tCategory\tSub-Category"]
                for rec in records:
                    qty = rec.get("QTY", 0)
                    name = rec.get("Item Name", "")
                    val = rec.get("Sell Value", "")
                    cat = rec.get("Category", "")
                    sub = rec.get("Sub-Category", "")
                    lines.append(f"{qty}\t{name}\t{val}\t{cat}\t{sub}")
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(lines))

            elif ext == ".pdf":
                # Build a simple HTML table and print to PDF
                html_rows = [
                    "<tr>"
                    "<th>QTY</th>"
                    "<th>Item Name</th>"
                    "<th>Sell Value</th>"
                    "<th>Category</th>"
                    "<th>Sub-Category</th>"
                    "</tr>"
                ]
                for rec in records:
                    qty = rec.get("QTY", 0)
                    name = rec.get("Item Name", "")
                    val = rec.get("Sell Value", "")
                    cat = rec.get("Category", "")
                    sub = rec.get("Sub-Category", "")
                    # Nicely format numeric sell values
                    if isinstance(val, (int, float)):
                        val_str = f"{val:,.0f}"
                    else:
                        val_str = "" if val is None else str(val)
                    html_rows.append(
                        "<tr>"
                        f"<td>{qty}</td>"
                        f"<td>{name}</td>"
                        f"<td>{val_str}</td>"
                        f"<td>{cat}</td>"
                        f"<td>{sub}</td>"
                        "</tr>"
                    )

                html = (
                    "<html><head><meta charset='utf-8'></head><body>"
                    "<h2>Inventory Index</h2>"
                    "<table border='1' cellspacing='0' cellpadding='3'>"
                    + "".join(html_rows) +
                    "</table>"
                    "</body></html>"
                )

                printer = QPrinter(QPrinter.HighResolution)
                printer.setOutputFormat(QPrinter.PdfFormat)
                printer.setOutputFileName(file_path)

                doc = QTextDocument()
                doc.setHtml(html)
                doc.print_(printer)

            QMessageBox.information(
                self,
                "Export Complete",
                f"Index exported to:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Failed",
                f"Failed to export index:\n{e}"
            )

    # ---------------- IMPORT INDEX (JSON, XLSX) ----------------
    def import_index(self):
        """
        Import index data from JSON or XLSX with fields:
        QTY, Item Name, Sell Value

        Matches items by Item Name to current market_data.
        Updates self.inventory and persists to INVENTORY_FILE.
        """
        file_path, selected_filter = QFileDialog.getOpenFileName(
            self,
            "Import Index",
            APP_DIR,
            "Index Files (*.json *.xlsx);;JSON (*.json);;Excel (*.xlsx)"
        )
        if not file_path:
            return  # user cancelled

        root, ext = os.path.splitext(file_path)
        ext = ext.lower()

        QApplication.setOverrideCursor(Qt.WaitCursor)
        self.status_label.setText("Importing index…")
        QApplication.processEvents()

        try:
            # Build quick lookup: item_name -> item dict
            name_map = {}
            for item in self.market_data:
                nm = item.get("item_name")
                if nm:
                    name_map[nm] = item

            records = []

            if ext == ".json":
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if not isinstance(data, list):
                    raise ValueError("JSON must be a list of objects.")
                # Normalize JSON records
                for rec in data:
                    if not isinstance(rec, dict):
                        continue
                    # Flexible key matching
                    name = (
                        rec.get("Item Name")
                        or rec.get("item_name")
                        or rec.get("name")
                    )
                    qty = (
                        rec.get("QTY")
                        if "QTY" in rec else rec.get("qty", 0)
                    )
                    sell_val = (
                        rec.get("Sell Value")
                        if "Sell Value" in rec
                        else rec.get("Sell Price", rec.get("sell_price"))
                    )
                    records.append({
                        "Item Name": name,
                        "QTY": qty,
                        "Sell Value": sell_val,
                    })

            elif ext == ".xlsx":
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active

                # Read header row and map columns
                header_map = {}
                first = True
                for row in ws.iter_rows(values_only=True):
                    if first:
                        first = False
                        if not row:
                            continue
                        for idx, val in enumerate(row):
                            if not val:
                                continue
                            key = str(val).strip().lower()
                            header_map[key] = idx
                        continue

                    if not row:
                        continue

                    def get_col(*names):
                        for n in names:
                            idx = header_map.get(n)
                            if idx is not None and idx < len(row):
                                return row[idx]
                        return None

                    name = get_col("item name", "name")
                    qty = get_col("qty", "quantity")
                    sell_val = get_col("sell value", "sell price")

                    records.append({
                        "Item Name": name,
                        "QTY": qty,
                        "Sell Value": sell_val,
                    })

            else:
                raise ValueError("Unsupported file type. Use .json or .xlsx")

            not_found = []

            for idx, rec in enumerate(records):
                if not isinstance(rec, dict):
                    continue

                if idx % 500 == 0:
                    QApplication.processEvents()

                name = rec.get("Item Name")
                if not name:
                    continue

                item = name_map.get(name)
                if not item:
                    not_found.append(name)
                    continue

                key = self.item_key(item)

                qty_raw = rec.get("QTY", 0)
                sell_value_raw = rec.get("Sell Value")

                # Parse qty
                try:
                    qty_val = float(qty_raw) if qty_raw is not None else 0.0
                except (TypeError, ValueError):
                    qty_val = 0.0

                # Start with existing record if any
                inv_entry = self.inventory.get(key)
                if isinstance(inv_entry, dict):
                    inv_entry = dict(inv_entry)
                else:
                    inv_entry = {}

                inv_entry["qty"] = qty_val

                # Parse sell_value -> stored as sell_price
                if sell_value_raw in (None, ""):
                    inv_entry.pop("sell_price", None)
                else:
                    try:
                        inv_entry["sell_price"] = float(sell_value_raw)
                    except (TypeError, ValueError):
                        inv_entry["sell_price"] = sell_value_raw

                # If both qty and sell_price are effectively empty, remove
                if qty_val == 0 and "sell_price" not in inv_entry:
                    self.inventory.pop(key, None)
                else:
                    self.inventory[key] = inv_entry

            # Persist to live inventory file
            save_inventory(INVENTORY_FILE, self.inventory)

            # Refresh UI
            self.apply_filters()
            self.update_overall_total()

            if not_found:
                skipped_list = "\n- ".join(sorted(set(not_found)))
                msg = (
                    "Import complete.\n\n"
                    "The following items from the file were not found in "
                    "the current market data and were skipped:\n"
                    f"- {skipped_list}"
                )
            else:
                msg = "Import complete."

            self.status_label.setText("Import complete.")
            QMessageBox.information(
                self,
                "Import Index",
                msg
            )

        except Exception as e:
            self.status_label.setText("Import failed.")
            QMessageBox.critical(
                self,
                "Import Failed",
                f"Failed to import index:\n{e}"
            )
        finally:
            QApplication.restoreOverrideCursor()

    # ------------------------------------------------------------


def apply_dynamic_font_scaling(app: QApplication):
    """
    Dynamically scale the base application font based on screen DPI,
    so it doesn't look tiny on high-DPI monitors.
    """
    screen = app.primaryScreen()
    if not screen:
        return

    dpi = screen.logicalDotsPerInch()  # typically 96 on standard displays
    scale = dpi / 96.0 if dpi else 1.0

    base_size = 9  # base point size
    font = app.font()
    font.setPointSizeF(base_size * scale)
    app.setFont(font)


def main():
    app = QApplication(sys.argv)

    # Dynamic font scaling based on DPI
    apply_dynamic_font_scaling(app)

    # Try to use the exe's embedded icon as the global app icon
    icon = get_exe_icon_qicon()
    if icon is not None:
        app.setWindowIcon(icon)

    win = AnnexWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
